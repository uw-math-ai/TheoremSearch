"""Export the ≥0.90 v2 consensus judge results to a review workbook.

Joins consensus_ge90_v2.jsonl (Opus 2-rater consensus verdicts) to
salient_matches_full.csv (formal/informal content) by (query_sid, cand_sid).

Produces consensus_ge90_v2_review.xlsx with 3 tabs:
  - Summary          : aggregate table COMPUTED from the row data (not hand-entered)
  - Verified Matches : edge==True  (confirmed + tiebroken + edge_ambiguous)
  - Verified Non-Matches : edge==False
Plus CSV fallbacks of each tab.

Column order puts the two slogans ADJACENT for fast eye-review, then the
three Opus ratings + reasons, then full formal code / informal statement.
"""
from __future__ import annotations
import csv, json
from collections import Counter
from pathlib import Path

HERE = Path(__file__).parent
# prefer the COMPLETE all-classes/sources file once present; else the mathlib→arXiv-only file
_ALL = HERE / "data" / "consensus_ge90_v2_all.jsonl"
CONS = _ALL if _ALL.exists() else HERE / "data" / "consensus_ge90_v2.jsonl"
CONTENT = HERE / "data" / "salient_matches_full.csv"
OUT_XLSX = HERE / "data" / "consensus_ge90_v2_review.xlsx"

# review-optimized column order: slogans adjacent, then ratings, then full text
COLUMNS = [
    ("sim", "similarity"),
    ("band", "sim_band"),
    ("project", "formal_project"),
    ("informal_source", "informal_source"),
    ("verdict", "verdict (edge?)"),
    ("final", "final_label"),
    ("status", "consensus_status"),
    ("formal_slogan", "FORMAL slogan"),
    ("informal_slogan", "INFORMAL slogan"),     # <- adjacent to formal slogan
    ("rater1", "opus_rater1"),
    ("rater2", "opus_rater2"),
    ("tiebreak", "opus_tiebreak"),
    ("reason1", "rater1_reason"),
    ("reason2", "rater2_reason"),
    ("reason3", "tiebreak_reason"),
    ("formal_decl", "formal_decl_name"),
    ("formal_body", "FORMAL code (Lean)"),
    ("informal_body", "INFORMAL statement (LaTeX)"),
    ("arxiv_id", "arxiv_id"),
    ("paper_title", "paper_title"),
    ("informal_ref", "informal_ref"),
    ("query_sid", "formal_statement_id"),
    ("cand_sid", "informal_statement_id"),
]


def load_joined():
    csv.field_size_limit(10_000_000)
    content = {}
    for r in csv.DictReader(CONTENT.open()):
        content[f'{r["query_sid"]}|{r["cand_sid"]}'] = r
    rows = []
    for l in CONS.open():
        l = l.strip()
        if not l:
            continue
        c = json.loads(l)
        ct = content.get(c["key"], {})
        labels = c.get("labels", [])
        reasons = c.get("reasons", [])
        isrc = (ct.get("informal_source", "") or "").strip()
        source = "arXiv" if "arx" in isrc.lower() else ("blueprint" if isrc else "")
        rows.append({
            "sim": float(c["sim"]), "band": c["band"],
            "project": ((ct.get("formal_module", "") or "").split(".")[0] or "(unknown)"),
            "informal_source": source,
            "verdict": "MATCH" if c["edge"] is True else ("NON-MATCH" if c["edge"] is False else "ambiguous"),
            "final": c["final"] if c["final"] else "(split)",
            "status": c["status"],
            "formal_slogan": ct.get("formal_slogan", ""),
            "informal_slogan": ct.get("informal_slogan", ""),
            "rater1": labels[0] if len(labels) > 0 else "",
            "rater2": labels[1] if len(labels) > 1 else "",
            "tiebreak": labels[2] if len(labels) > 2 else "",
            "reason1": reasons[0] if len(reasons) > 0 else "",
            "reason2": reasons[1] if len(reasons) > 1 else "",
            "reason3": reasons[2] if len(reasons) > 2 else "",
            "formal_decl": c["formal_decl"], "formal_body": ct.get("formal_body", ""),
            "informal_body": ct.get("informal_body", ""),
            "arxiv_id": c["arxiv_id"], "paper_title": ct.get("paper_title", ""),
            "informal_ref": ct.get("informal_ref", ""),
            "query_sid": c["query_sid"], "cand_sid": c["cand_sid"],
            "_edge": c["edge"],
        })
    return rows


def _rate(m, t):
    return f"{100*m/t:.1f}%" if t else ""


def aggregate(rows):
    """Aggregate tables COMPUTED from the row data — one value per cell, with a
    real header row per section (first column header doubles as the section
    name). Used for both the Summary tab and consensus_ge90_v2_summary.csv."""
    n = len(rows)
    edge = Counter(r["_edge"] for r in rows)
    matches = [r for r in rows if r["_edge"] is True]
    out = []

    def grouped(header, keyfn):
        """Emit a match/no-match/rate table grouped by keyfn, sorted by size."""
        out.append([header, "candidates", "matched", "no-match", "match rate"])
        groups = {}
        for r in rows:
            groups.setdefault(keyfn(r), []).append(r)
        for k in sorted(groups, key=lambda p: -len(groups[p])):
            sub = groups[k]
            m = sum(1 for r in sub if r["_edge"] is True)
            nm = sum(1 for r in sub if r["_edge"] is False)
            out.append([k, len(sub), m, nm, _rate(m, len(sub))])

    out.append(["≥0.90 candidate-edge judge results (all classes/sources) — v2 guarded prompt, 2-rater Opus consensus + tie-break"])
    out.append([])
    out.append(['"matched" = same theorem (exact OR inexact); "no-match" = different theorem (wrong)'])
    out.append([])

    # --- per band: matched / no-match / rate ---
    out.append(["by similarity band", "candidates", "matched", "no-match", "match rate"])
    for band in ("0.95-1.0", "0.90-0.95"):
        sub = [r for r in rows if r["band"] == band]
        m = sum(1 for r in sub if r["_edge"] is True)
        nm = sum(1 for r in sub if r["_edge"] is False)
        out.append([band, len(sub), m, nm, _rate(m, len(sub))])
    out.append(["TOTAL (≥0.90)", n, edge[True], edge[False], _rate(edge[True], n)])
    out.append([])

    # --- match strength of the verified matches: one count per cell ---
    fin = Counter(r["final"] for r in matches)
    nm_ = len(matches)
    out.append(["match strength (of matches)", "count", "share of matches"])
    for label, key in (("exact", "exact"), ("inexact", "inexact"), ("split", "(split)")):
        out.append([label, fin.get(key, 0), _rate(fin.get(key, 0), nm_)])
    out.append([])

    # --- consensus agreement: one count per cell ---
    st = Counter(r["status"] for r in rows)
    out.append(["consensus agreement", "count", "share of all"])
    out.append(["confirmed (2 raters agreed)", st.get("confirmed", 0), _rate(st.get("confirmed", 0), n)])
    out.append(["tie-broken (needed 3rd rater)", st.get("tiebroken", 0), _rate(st.get("tiebroken", 0), n)])
    out.append(["edge-ambiguous (agree edge, split strength)", st.get("edge_ambiguous", 0), _rate(st.get("edge_ambiguous", 0), n)])
    out.append([])

    # --- per-band strength breakdown: one count per cell (exact+inexact+split+wrong = candidates) ---
    out.append(["band strength breakdown", "candidates", "exact", "inexact", "split", "wrong"])
    for band in ("0.95-1.0", "0.90-0.95"):
        sub = [r for r in rows if r["band"] == band]
        if not sub:
            continue
        ex = sum(1 for r in sub if r["final"] == "exact")
        ix = sum(1 for r in sub if r["final"] == "inexact")
        wr = sum(1 for r in sub if r["final"] == "wrong")
        sp = sum(1 for r in sub if r["final"] == "(split)")
        out.append([band, len(sub), ex, ix, sp, wr])
    out.append([])

    # --- per informal source and per formal project ---
    grouped("by informal source", lambda r: r["informal_source"] or "(none)")
    out.append([])
    grouped("by formal project", lambda r: r["project"])
    return out


def add_formula_aggregates(wb, rows):
    """A live-formula 'Aggregates' tab: every count is a COUNTIF against the
    Verified Matches / Verified Non-Matches tabs (not a Python-baked value), so
    the numbers recompute in Excel if you hand-edit a verdict during review and
    every cell is auditable. Column letters: B=sim_band, C=formal_project,
    D=informal_source, F=final_label (see COLUMNS)."""
    from openpyxl.styles import Font, PatternFill
    M = "'Verified Matches'"
    N = "'Verified Non-Matches'"
    s = wb.create_sheet("Aggregates", index=1)
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="1F4E78"); hdr_font = Font(bold=True, color="FFFFFF")
    r = 1

    def title(t):
        nonlocal r
        s.cell(r, 1, t).font = Font(bold=True, size=12); r += 2

    def section(label, col_letter, values):
        """label/col_letter = which column in the tabs to group by; values = row labels."""
        nonlocal r
        s.cell(r, 1, label).font = bold; r += 1
        for j, h in enumerate(["", "matched", "no-match", "total", "match rate"], 1):
            c = s.cell(r, j, h); c.fill = hdr_fill; c.font = hdr_font
        r += 1
        start = r
        for v in values:
            s.cell(r, 1, v)
            s.cell(r, 2, f"=COUNTIF({M}!${col_letter}:${col_letter},$A{r})")
            s.cell(r, 3, f"=COUNTIF({N}!${col_letter}:${col_letter},$A{r})")
            s.cell(r, 4, f"=B{r}+C{r}")
            mr = s.cell(r, 5, f"=IF(D{r}=0,\"\",B{r}/D{r})"); mr.number_format = "0.0%"
            r += 1
        # subtotal row
        s.cell(r, 1, "TOTAL").font = bold
        s.cell(r, 2, f"=SUM(B{start}:B{r-1})").font = bold
        s.cell(r, 3, f"=SUM(C{start}:C{r-1})").font = bold
        s.cell(r, 4, f"=B{r}+C{r}").font = bold
        tm = s.cell(r, 5, f"=IF(D{r}=0,\"\",B{r}/D{r})"); tm.number_format = "0.0%"; tm.font = bold
        r += 2

    title("≥0.90 candidate-edge aggregates — LIVE FORMULAS over the verified tabs")
    s.cell(r, 1, "Every count = COUNTIF against 'Verified Matches' / 'Verified Non-Matches'; edit a verdict there and these recompute.")
    r += 2

    section("— by similarity band —", "B", ["0.95-1.0", "0.90-0.95"])
    projects = sorted({x["project"] for x in rows},
                      key=lambda p: -sum(1 for x in rows if x["project"] == p))
    section("— by formal project —", "C", projects)
    sources = sorted({(x["informal_source"] or "(none)") for x in rows},
                     key=lambda p: -sum(1 for x in rows if (x["informal_source"] or "(none)") == p))
    section("— by informal source —", "D", sources)

    # match strength (only meaningful on the matches tab): exact / inexact / split
    s.cell(r, 1, "— match strength (Verified Matches tab) —").font = bold; r += 1
    for lab, key in [("exact", "exact"), ("inexact", "inexact"), ("(split)", "(split)")]:
        s.cell(r, 1, lab)
        s.cell(r, 2, f"=COUNTIF({M}!$F:$F,$A{r})")
        pc = s.cell(r, 3, f"=IF(COUNTA({M}!$F:$F)-1=0,\"\",B{r}/(COUNTA({M}!$F:$F)-1))")
        pc.number_format = "0.0%"
        r += 1

    s.column_dimensions["A"].width = 34
    for col in ("B", "C", "D", "E"):
        s.column_dimensions[col].width = 12
    return s


def main():
    rows = load_joined()
    matches = [r for r in rows if r["_edge"] is True]
    nonm = [r for r in rows if r["_edge"] is False]
    # sort each: similarity desc (best first)
    matches.sort(key=lambda r: -r["sim"])
    nonm.sort(key=lambda r: -r["sim"])

    # assertions: no data loss, counts reconcile (7,217 mathlib→arXiv, or ~8,022 complete)
    assert len(rows) >= 7000, f"too few rows: {len(rows)}"
    amb = sum(1 for r in rows if r["_edge"] is None)
    assert len(matches) + len(nonm) + amb == len(rows), "counts don't reconcile"
    assert all(r["formal_slogan"] or r["formal_body"] for r in rows[:50]), "content join looks empty"
    print(f"input: {CONS.name} ({len(rows)} rows)")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()

    # --- Summary ---
    ws = wb.active; ws.title = "Summary"
    for row in aggregate(rows):
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=12)
    # bold each section's header row (identified by its 2nd-column header token)
    _hdr_tokens = {"candidates", "count"}
    for row in ws.iter_rows(min_row=2):
        if len(row) > 1 and row[1].value in _hdr_tokens:
            for cell in row:
                if cell.value not in (None, ""):
                    cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 42
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 12

    # live-formula aggregates tab (auditable cross-check of the Python Summary)
    add_formula_aggregates(wb, rows)

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")

    def write_sheet(title, data):
        s = wb.create_sheet(title)
        s.append([disp for _, disp in COLUMNS])
        for c in range(1, len(COLUMNS) + 1):
            cell = s.cell(row=1, column=c); cell.fill = hdr_fill; cell.font = hdr_font
        for r in data:
            s.append([round(r[k], 4) if k == "sim" else r.get(k, "") for k, _ in COLUMNS])
        s.freeze_panes = "A2"
        # widths tuned for review: slogans + bodies wide, labels narrow
        widths = {"FORMAL slogan": 55, "INFORMAL slogan": 55, "FORMAL code (Lean)": 60,
                  "INFORMAL statement (LaTeX)": 60, "rater1_reason": 50, "rater2_reason": 50,
                  "tiebreak_reason": 50, "formal_decl_name": 34, "paper_title": 34}
        from openpyxl.utils import get_column_letter
        for i, (_, disp) in enumerate(COLUMNS, 1):
            s.column_dimensions[get_column_letter(i)].width = widths.get(disp, 13)
        # wrap the long text columns
        wrapcols = {i for i, (_, d) in enumerate(COLUMNS, 1)
                    if d in ("FORMAL slogan", "INFORMAL slogan", "FORMAL code (Lean)",
                             "INFORMAL statement (LaTeX)", "rater1_reason", "rater2_reason", "tiebreak_reason")}
        for row in s.iter_rows(min_row=2):
            for cell in row:
                if cell.column in wrapcols:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    write_sheet("Verified Matches", matches)
    write_sheet("Verified Non-Matches", nonm)
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"  Verified Matches:     {len(matches):,} rows")
    print(f"  Verified Non-Matches: {len(nonm):,} rows")
    print(f"  ambiguous (excluded from both tabs): {amb}")

    # CSV fallbacks
    for name, data in [("matches", matches), ("non_matches", nonm)]:
        p = HERE / "data" / f"consensus_ge90_v2_{name}.csv"
        with p.open("w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow([disp for _, disp in COLUMNS])
            for r in data:
                w.writerow([r.get(k, "") for k, _ in COLUMNS])
        print(f"  csv: {p.name} ({len(data):,} rows)")
    # aggregate csv
    p = HERE / "data" / "consensus_ge90_v2_summary.csv"
    with p.open("w", newline="") as fh:
        csv.writer(fh).writerows(aggregate(rows))
    print(f"  csv: {p.name}")


if __name__ == "__main__":
    main()
